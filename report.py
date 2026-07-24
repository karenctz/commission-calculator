"""Builds an .xlsx matching the layout of the real
`Karen - Commission - Payment on 26 Jul 26.xlsx`: a company/title/legend
header block, one flat chronological listing with a bold PO-ref divider
row only when a supplier PO exists (no subtotal per group - matches the
real file, which only subtotals once at the end), unpaid invoices in red,
and a single grand total row. A "Payout by salesperson" appendix (not in
the original file, but requested separately) follows the total. Ignored
invoices are excluded from every total but listed in a separate sheet for
audit trail.
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import commission

HEADERS = ["No.", "Posting Date", "Customer", "Amount", "Cost", "GP", "Comm %", "Comm Amt", "Remarks"]
BOLD = Font(bold=True)
RED = Font(color="FFFF0000")
PO_HEADER_FONT = Font(bold=True, color="FF0070C0")
WHT_FILL = PatternFill(fill_type="solid", fgColor="FFFFE699")


def _comm_pct_display(line_items_df, invoice_no):
    pcts = line_items_df.loc[line_items_df["invoice_no"] == invoice_no, "commission_pct"].unique()
    if len(pcts) == 1:
        return pcts[0]
    return "mixed"


def build_report(invoices_df, line_items_df, require_approved=True, require_paid=True,
                  salesperson=None, date_from=None, date_to=None):
    """Payout eligibility defaults to commission_approved AND paid_by_customer
    AND NOT ignored (see plan) - the two require_* flags let Finance loosen
    this deliberately (e.g. to preview approved-but-unpaid amounts) rather
    than the report silently including invoices that aren't actually payable
    yet."""
    invoices = invoices_df[~invoices_df["ignored"]].copy()
    if require_approved:
        invoices = invoices[invoices["commission_approved"]]
    if require_paid:
        invoices = invoices[invoices["paid_by_customer"]]
    if salesperson and salesperson != "All":
        invoices = invoices[invoices["salesperson"] == salesperson]
    if date_from:
        invoices = invoices[invoices["invoice_date"] >= date_from]
    if date_to:
        invoices = invoices[invoices["invoice_date"] <= date_to]
    invoices = invoices.sort_values("invoice_date")

    year = max((d[:4] for d in invoices["invoice_date"] if d), default=str(date.today().year))
    title = (
        f"{salesperson} Commission Listing {year}"
        if salesperson and salesperson != "All"
        else f"Commission Listing {year}"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Commission Report"

    ws.append([
        "CACTOZ PTE LTD", None, title, None, "Invoice in RED = Payment outstanding",
        None, None, None, None, None,
        "Infrastructure and Hardware Professional Service = 30% as cost, 70% as profit",
    ])
    ws["A1"].font = BOLD
    ws["C1"].font = BOLD

    ws.append([
        None, None, None, None, "Document missing", None, None, "Please advise PS Costing",
        None, None, "Apps Team Professional Service = 70% as cost, 30% as profit",
    ])

    ws.append(HEADERS)
    for cell in ws[ws.max_row]:
        cell.font = BOLD

    grand_amount = grand_cost = grand_gp = grand_comm = 0.0
    prev_po = None

    for invoice_no, inv in invoices.iterrows():
        po_ref = (inv.get("po_ref") or "").strip()
        if po_ref and po_ref != prev_po:
            ws.append([po_ref])
            ws[ws.max_row][0].font = PO_HEADER_FONT
        if po_ref:
            prev_po = po_ref

        rollup = commission.invoice_rollup(line_items_df, invoice_no)
        comm_pct = _comm_pct_display(line_items_df, invoice_no)
        remarks = inv.get("notes") or ""
        if not inv.get("invoice_pdf_path"):
            remarks = f"{remarks} (document missing)".strip()
        if rollup.get("wht_total"):
            remarks = f"{remarks} (WHT ${rollup['wht_total']:,.2f} deducted - verify)".strip()
        row = [
            invoice_no, inv["invoice_date"], inv["customer"],
            rollup["selling_total"], rollup["cost_total"], rollup["margin_total"],
            comm_pct, rollup["commission_total"], remarks,
        ]
        ws.append(row)
        if not inv["paid_by_customer"]:
            for cell in ws[ws.max_row]:
                cell.font = RED
        if rollup.get("wht_total"):
            for cell in ws[ws.max_row]:
                cell.fill = WHT_FILL

        grand_amount += rollup["selling_total"]
        grand_cost += rollup["cost_total"]
        grand_gp += rollup["margin_total"]
        grand_comm += rollup["commission_total"]

    ws.append([
        None, None, "Total",
        round(grand_amount, 2), round(grand_cost, 2), round(grand_gp, 2),
        None, round(grand_comm, 2),
    ])
    for cell in ws[ws.max_row]:
        cell.font = BOLD

    if len(invoices):
        ws.append([])
        ws.append(["Payout by salesperson"])
        ws[ws.max_row][0].font = BOLD
        for sp, sp_group in invoices.groupby("salesperson"):
            sp_commission = sum(
                commission.invoice_rollup(line_items_df, no)["commission_total"] for no in sp_group["invoice_no"]
            )
            ws.append([sp, None, None, None, None, None, None, round(sp_commission, 2)])

    ignored = invoices_df[invoices_df["ignored"]]
    if len(ignored):
        ws2 = wb.create_sheet("Ignored (excluded)")
        ws2.append(HEADERS[:3] + ["Reason"])
        for cell in ws2[1]:
            cell.font = BOLD
        for invoice_no, inv in ignored.iterrows():
            ws2.append([invoice_no, inv["invoice_date"], inv["customer"], inv["ignore_reason"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
